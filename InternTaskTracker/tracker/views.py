import csv
from datetime import timedelta

from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.paginator import Paginator
from django.db.models import Count, DurationField, Q, Sum, Value
from django.db.models.functions import Coalesce
from django.http import FileResponse, Http404, HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from .forms import (
    AcceptInviteForm,
    AssignInternForm,
    AssignTaskForm,
    CategoryForm,
    InviteUserForm,
    InternTaskForm,
    MessageForm,
    ProfileUpdateForm,
    StartChatForm,
    TaskReviewForm,
    TaskUpdateForm,
    UserUpdateForm,
)
from .invites import create_invitation, invitation_absolute_url, send_invitation_email
from .models import (
    Category, Conversation, Invitation, Message, Notification, Task, TaskReview, User,
)
from .permissions import (
    can_delete_task,
    can_edit_task,
    can_message,
    can_review_task,
    can_submit_task,
    can_view_task,
    categories_for,
    messageable_users_for,
    visible_interns_for,
    visible_tasks_for,
)

ZERO = Value(timedelta(), output_field=DurationField())


def is_head(user):
    return user.is_authenticated and user.is_head


def is_management(user):
    return user.is_authenticated and user.is_management


head_required = user_passes_test(is_head)
management_required = user_passes_test(is_management)


def hours(duration):
    if not duration:
        return 0
    return round(duration.total_seconds() / 3600, 1)


def notify(user, title, message="", link=""):
    Notification.objects.create(user=user, title=title, message=message, link=link)


# ---------------------------------------------------------------- public landing

def landing(request):
    if request.user.is_authenticated:
        return redirect("dashboard")
    return render(request, "landing.html")


# ---------------------------------------------------------------- auth / invites

def accept_invite(request, token):
    invitation = get_object_or_404(Invitation, token=token)
    if invitation.is_accepted:
        messages.info(request, "This invitation has already been used. Please log in.")
        return redirect("login")
    if invitation.is_expired:
        messages.error(request, "This invitation link has expired. Ask your department head to resend it.")
        return redirect("login")

    user = invitation.user
    if request.method == "POST":
        form = AcceptInviteForm(user, request.POST)
        if form.is_valid():
            form.save()
            invitation.accepted_at = timezone.now()
            invitation.save(update_fields=["accepted_at"])
            login(request, user)
            messages.success(request, "Email verified. Welcome to InTrack!")
            return redirect("dashboard")
    else:
        form = AcceptInviteForm(user)

    return render(request, "registration/accept_invite.html", {
        "form": form,
        "invite_user": user,
    })


# ---------------------------------------------------------------- helpers

def filter_tasks(request, queryset):
    params = request.GET
    if q := params.get("q", "").strip():
        queryset = queryset.filter(
            Q(title__icontains=q) | Q(description__icontains=q)
        )
    if status := params.get("status"):
        queryset = queryset.filter(status=status)
    if priority := params.get("priority"):
        queryset = queryset.filter(priority=priority)
    if category := params.get("category"):
        queryset = queryset.filter(category_id=category)
    if date := params.get("date"):
        queryset = queryset.filter(date=date)
    if date_from := params.get("date_from"):
        queryset = queryset.filter(date__gte=date_from)
    if date_to := params.get("date_to"):
        queryset = queryset.filter(date__lte=date_to)
    if month := params.get("month"):
        try:
            year, mon = month.split("-")
            queryset = queryset.filter(date__year=int(year), date__month=int(mon))
        except ValueError:
            pass
    if intern := params.get("intern"):
        queryset = queryset.filter(intern_id=intern)
    return queryset


def week_bounds(today):
    start = today - timedelta(days=today.weekday())
    return start, start + timedelta(days=6)


def weekly_chart_data(queryset, today):
    week_start, _ = week_bounds(today)
    labels, data = [], []
    for i in range(7):
        day = week_start + timedelta(days=i)
        total = queryset.filter(date=day).aggregate(
            t=Coalesce(Sum("duration"), ZERO))["t"]
        labels.append(day.strftime("%a"))
        data.append(hours(total))
    return labels, data


# ---------------------------------------------------------------- dashboard

@login_required
def dashboard(request):
    if request.user.is_head:
        return management_dashboard(request, scope="department")
    if request.user.is_supervisor:
        return management_dashboard(request, scope="team")
    return intern_dashboard(request)


def intern_dashboard(request):
    today = timezone.localdate()
    week_start, week_end = week_bounds(today)
    tasks = request.user.tasks.all()

    stats = {
        "hours_today": hours(tasks.filter(date=today).aggregate(
            t=Coalesce(Sum("duration"), ZERO))["t"]),
        "hours_week": hours(tasks.filter(date__range=(week_start, week_end)).aggregate(
            t=Coalesce(Sum("duration"), ZERO))["t"]),
        "hours_month": hours(tasks.filter(
            date__year=today.year, date__month=today.month).aggregate(
            t=Coalesce(Sum("duration"), ZERO))["t"]),
        "completed": tasks.filter(status=Task.Status.APPROVED).count(),
        "pending": tasks.filter(status=Task.Status.PENDING).count(),
        "in_progress": tasks.filter(status=Task.Status.IN_PROGRESS).count(),
        "changes_requested": tasks.filter(status=Task.Status.CHANGES_REQUESTED).count(),
        "submitted": tasks.filter(status=Task.Status.SUBMITTED).count(),
    }
    chart_labels, chart_data = weekly_chart_data(tasks, today)

    return render(request, "dashboard/intern_dashboard.html", {
        "stats": stats,
        "recent_tasks": tasks.select_related("category", "assigned_by")[:6],
        "chart_labels": chart_labels,
        "chart_data": chart_data,
        "assigned_supervisor": request.user.supervisor,
    })


def management_dashboard(request, scope="team"):
    today = timezone.localdate()
    tasks = visible_tasks_for(request.user)
    interns = visible_interns_for(request.user)

    stats = {
        "total_interns": interns.count(),
        "total_tasks": tasks.count(),
        "active_tasks": tasks.filter(
            status__in=[Task.Status.PENDING, Task.Status.IN_PROGRESS,
                        Task.Status.CHANGES_REQUESTED]).count(),
        "awaiting_review": tasks.filter(status=Task.Status.SUBMITTED).count(),
        "total_hours": hours(tasks.aggregate(t=Coalesce(Sum("duration"), ZERO))["t"]),
    }

    top_interns = (
        interns.annotate(
            total_hours=Coalesce(Sum("tasks__duration"), ZERO),
            task_count=Count("tasks"),
        )
        .order_by("-total_hours")[:5]
    )
    top_interns = [
        {"user": u, "hours": hours(u.total_hours), "task_count": u.task_count}
        for u in top_interns
    ]
    chart_labels, chart_data = weekly_chart_data(tasks, today)

    title = "Department Dashboard" if scope == "department" else "Supervisor Dashboard"
    return render(request, "dashboard/supervisor_dashboard.html", {
        "stats": stats,
        "top_interns": top_interns,
        "recent_tasks": tasks.select_related("intern", "category")[:8],
        "chart_labels": chart_labels,
        "chart_data": chart_data,
        "page_heading": title,
        "scope": scope,
    })


# ---------------------------------------------------------------- tasks

@login_required
def task_list(request):
    tasks = filter_tasks(request, visible_tasks_for(request.user))
    paginator = Paginator(tasks, 10)
    page = paginator.get_page(request.GET.get("page"))
    total_hours = hours(tasks.aggregate(t=Coalesce(Sum("duration"), ZERO))["t"])

    return render(request, "tasks/task_list.html", {
        "page": page,
        "total_hours": total_hours,
        "categories": categories_for(request.user),
        "interns": visible_interns_for(request.user) if request.user.is_management else None,
        "statuses": Task.Status.choices,
        "priorities": Task.Priority.choices,
    })


@login_required
def task_detail(request, pk):
    task = get_object_or_404(
        Task.objects.select_related("intern", "category", "assigned_by"), pk=pk)
    if not can_view_task(request.user, task):
        messages.error(request, "You do not have permission to view this task.")
        return redirect("task_list")
    review_form = TaskReviewForm() if can_review_task(request.user, task) else None
    return render(request, "tasks/task_detail.html", {
        "task": task,
        "can_edit": can_edit_task(request.user, task),
        "can_delete": can_delete_task(request.user, task),
        "can_submit": can_submit_task(request.user, task),
        "can_review": can_review_task(request.user, task),
        "review_form": review_form,
        "reviews": task.reviews.select_related("reviewer"),
    })


@login_required
@require_POST
def task_submit(request, pk):
    task = get_object_or_404(Task, pk=pk)
    if not can_submit_task(request.user, task):
        messages.error(request, "You cannot submit this task for review.")
        return redirect("task_detail", pk=pk)
    task.status = Task.Status.SUBMITTED
    task.save(update_fields=["status", "updated_at"])

    reviewers = []
    if task.intern.supervisor_id:
        reviewers.append(task.intern.supervisor)
    head = User.objects.filter(
        role=User.Roles.HEAD, department=task.intern.department).first()
    if head and head not in reviewers:
        reviewers.append(head)
    for reviewer in reviewers:
        notify(reviewer, "Task submitted for review",
               f"{request.user} submitted \"{task.title}\" for review.",
               link=f"/tasks/{task.pk}/")

    messages.success(request, "Task submitted for review.")
    return redirect("task_detail", pk=pk)


@login_required
@require_POST
def task_review(request, pk):
    task = get_object_or_404(Task, pk=pk)
    if not can_review_task(request.user, task):
        messages.error(request, "You cannot review this task.")
        return redirect("task_detail", pk=pk)

    form = TaskReviewForm(request.POST)
    if not form.is_valid():
        messages.error(request, "Please fix the review form errors.")
        return redirect("task_detail", pk=pk)

    action = form.cleaned_data["action"]
    comment = (form.cleaned_data.get("comment") or "").strip()
    if action == "APPROVE":
        task.status = Task.Status.APPROVED
        review_action = TaskReview.Action.APPROVE
        notify(task.intern, "Task approved",
               f"{request.user} approved \"{task.title}\".",
               link=f"/tasks/{task.pk}/")
        messages.success(request, "Task approved.")
    else:
        task.status = Task.Status.CHANGES_REQUESTED
        review_action = TaskReview.Action.REQUEST_CHANGES
        notify(task.intern, "Changes requested",
               f"{request.user} requested changes on \"{task.title}\": {comment}",
               link=f"/tasks/{task.pk}/")
        messages.success(request, "Changes requested. The intern has been notified.")

    task.save(update_fields=["status", "updated_at"])
    TaskReview.objects.create(
        task=task, reviewer=request.user, action=review_action, comment=comment)
    return redirect("task_detail", pk=pk)


@login_required
def task_attachment(request, pk):
    task = get_object_or_404(Task, pk=pk)
    if not can_view_task(request.user, task):
        return HttpResponseForbidden("Not allowed.")
    if not task.attachment:
        raise Http404("No attachment.")
    return FileResponse(
        task.attachment.open("rb"),
        as_attachment=True,
        filename=task.attachment.name.split("/")[-1],
    )


@login_required
def task_create(request):
    """Interns log personal work; supervisors/heads assign tasks."""
    user = request.user
    if user.is_intern:
        form_class = InternTaskForm
        title = "Log Work"
    elif user.is_management:
        form_class = AssignTaskForm
        title = "Assign Task"
    else:
        return HttpResponseForbidden("Not allowed.")

    if request.method == "POST":
        form = form_class(request.POST, request.FILES, user=user)
        if form.is_valid():
            task = form.save(commit=False)
            if user.is_intern:
                task.intern = user
                task.assigned_by = None
            else:
                task.assigned_by = user
            task.save()

            if user.is_intern and user.supervisor:
                notify(user.supervisor, "New work logged",
                       f"{user} logged \"{task.title}\".",
                       link=f"/tasks/{task.pk}/")
            elif user.is_management:
                notify(task.intern, "New task assigned",
                       f"{user} assigned \"{task.title}\" to you.",
                       link=f"/tasks/{task.pk}/")

            messages.success(request, "Task saved successfully.")
            return redirect("task_list")
    else:
        form = form_class(user=user, initial={"date": timezone.localdate()})

    return render(request, "tasks/task_form.html", {"form": form, "title": title})


@login_required
def task_update(request, pk):
    task = get_object_or_404(Task, pk=pk)
    if not can_edit_task(request.user, task):
        messages.error(request, "You cannot edit this task.")
        return redirect("task_list")

    if request.method == "POST":
        form = TaskUpdateForm(
            request.POST, request.FILES, instance=task,
            user=request.user, assigned=task.is_assigned)
        if form.is_valid():
            form.save()
            messages.success(request, "Task updated successfully.")
            return redirect("task_detail", pk=task.pk)
    else:
        form = TaskUpdateForm(
            instance=task, user=request.user, assigned=task.is_assigned)

    return render(request, "tasks/task_form.html", {"form": form, "title": "Edit Task"})


@login_required
def task_delete(request, pk):
    task = get_object_or_404(Task, pk=pk)
    if not can_delete_task(request.user, task):
        messages.error(request, "You cannot delete this task.")
        return redirect("task_list")
    if request.method == "POST":
        task.delete()
        messages.success(request, "Task deleted.")
        return redirect("task_list")
    return render(request, "tasks/task_confirm_delete.html", {"task": task})


# ---------------------------------------------------------------- team (head)

@login_required
@head_required
def team_list(request):
    dept = request.user.department
    members = User.objects.filter(department=dept).exclude(pk=request.user.pk).select_related(
        "supervisor", "profile", "invitation").order_by("role", "first_name")
    if q := request.GET.get("q", "").strip():
        members = members.filter(
            Q(first_name__icontains=q) | Q(last_name__icontains=q) |
            Q(username__icontains=q) | Q(email__icontains=q)
        )
    member_rows = []
    for member in members:
        invite_link = ""
        inv = Invitation.objects.filter(user=member).first()
        if inv and not member.email_verified and not inv.is_accepted and not inv.is_expired:
            invite_link = invitation_absolute_url(request, inv)
        member_rows.append({"user": member, "invite_link": invite_link})
    return render(request, "team/team_list.html", {
        "member_rows": member_rows,
        "department": dept,
    })


@login_required
@head_required
def invite_user(request):
    if not request.user.department_id:
        messages.error(request, "Your account is not linked to a department. Contact the system admin.")
        return redirect("dashboard")

    if request.method == "POST":
        form = InviteUserForm(request.POST, head=request.user)
        if form.is_valid():
            user = form.save()
            invitation = create_invitation(user, invited_by=request.user)
            link = invitation_absolute_url(request, invitation)
            try:
                send_invitation_email(request, invitation)
                messages.success(
                    request,
                    f"{user.get_role_display()} {user} created. "
                    f"Invite emailed to {user.email}. Activation link: {link}"
                )
            except Exception:
                messages.warning(
                    request,
                    f"{user} was created, but the invite email could not be sent. "
                    f"Share this activation link: {link}"
                )
            return redirect("team_list")
    else:
        form = InviteUserForm(head=request.user)

    return render(request, "team/invite_user.html", {"form": form})


@login_required
@head_required
@require_POST
def resend_invite(request, pk):
    member = get_object_or_404(User, pk=pk, department=request.user.department)
    if member.email_verified and not member.must_set_password:
        messages.info(request, f"{member} has already activated their account.")
        return redirect("team_list")
    invitation = create_invitation(member, invited_by=request.user)
    link = invitation_absolute_url(request, invitation)
    try:
        send_invitation_email(request, invitation)
        messages.success(
            request,
            f"Invitation resent to {member.email}. Activation link: {link}"
        )
    except Exception:
        messages.warning(
            request,
            f"Email could not be sent. Share this activation link: {link}"
        )
    return redirect("team_list")


@login_required
@head_required
def category_list(request):
    categories = Category.objects.filter(department=request.user.department)
    return render(request, "team/category_list.html", {"categories": categories})


@login_required
@head_required
def category_create(request):
    if request.method == "POST":
        form = CategoryForm(request.POST)
        if form.is_valid():
            category = form.save(commit=False)
            category.department = request.user.department
            category.save()
            messages.success(request, f'Category "{category.name}" created.')
            return redirect("category_list")
    else:
        form = CategoryForm()
    return render(request, "team/category_form.html", {
        "form": form, "title": "Add Category"})


@login_required
@head_required
def category_edit(request, pk):
    category = get_object_or_404(
        Category, pk=pk, department=request.user.department)
    if request.method == "POST":
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, "Category updated.")
            return redirect("category_list")
    else:
        form = CategoryForm(instance=category)
    return render(request, "team/category_form.html", {
        "form": form, "title": "Edit Category"})


@login_required
@head_required
def assign_intern(request):
    if request.method == "POST":
        form = AssignInternForm(request.POST, head=request.user)
        if form.is_valid():
            intern = form.cleaned_data["intern"]
            supervisor = form.cleaned_data["supervisor"]
            intern.supervisor = supervisor
            intern.save(update_fields=["supervisor"])
            notify(intern, "Supervisor assigned",
                   f"You have been assigned to {supervisor}.",
                   link="/profile/")
            notify(supervisor, "New intern assigned",
                   f"{intern} has been assigned to you.",
                   link="/interns/")
            messages.success(request, f"{intern} assigned to {supervisor}.")
            return redirect("team_list")
    else:
        form = AssignInternForm(head=request.user)
    return render(request, "team/assign_intern.html", {"form": form})


# ---------------------------------------------------------------- supervisor / head shared

@login_required
@management_required
def intern_list(request):
    interns = visible_interns_for(request.user).select_related("profile", "supervisor")
    if q := request.GET.get("q", "").strip():
        interns = interns.filter(
            Q(first_name__icontains=q) | Q(last_name__icontains=q) |
            Q(username__icontains=q) | Q(email__icontains=q)
        )
    interns = interns.annotate(
        total_hours=Coalesce(Sum("tasks__duration"), ZERO),
        task_count=Count("tasks"),
        completed_count=Count(
            "tasks", filter=Q(tasks__status=Task.Status.APPROVED)),
    )
    rows = [
        {"user": u, "hours": hours(u.total_hours), "task_count": u.task_count,
         "completed_count": u.completed_count}
        for u in interns
    ]
    return render(request, "supervisor/intern_list.html", {"rows": rows})


@login_required
@management_required
def analytics(request):
    today = timezone.localdate()
    tasks = visible_tasks_for(request.user)
    interns = visible_interns_for(request.user)

    week_labels, week_data = [], []
    this_week_start, _ = week_bounds(today)
    for i in range(7, -1, -1):
        start = this_week_start - timedelta(weeks=i)
        end = start + timedelta(days=6)
        total = tasks.filter(date__range=(start, end)).aggregate(
            t=Coalesce(Sum("duration"), ZERO))["t"]
        week_labels.append(start.strftime("%d %b"))
        week_data.append(hours(total))

    month_labels, month_data = [], []
    year, month = today.year, today.month
    months = []
    for _ in range(6):
        months.append((year, month))
        month -= 1
        if month == 0:
            year, month = year - 1, 12
    for y, m in reversed(months):
        total = tasks.filter(date__year=y, date__month=m).aggregate(
            t=Coalesce(Sum("duration"), ZERO))["t"]
        month_labels.append(f"{m:02d}/{y}")
        month_data.append(hours(total))

    status_counts = {
        label: tasks.filter(status=value).count()
        for value, label in Task.Status.choices
    }

    intern_rows = interns.annotate(
        total_hours=Coalesce(Sum("tasks__duration"), ZERO)
    ).order_by("-total_hours")
    intern_labels = [u.get_full_name() or u.username for u in intern_rows]
    intern_data = [hours(u.total_hours) for u in intern_rows]

    category_rows = (
        Category.objects.filter(tasks__in=tasks)
        .annotate(n=Count("tasks", distinct=True))
        .filter(n__gt=0)
        .order_by("-n")
    )
    uncategorized = tasks.filter(category__isnull=True).count()
    category_labels = [c.name for c in category_rows]
    category_data = [c.n for c in category_rows]
    if uncategorized:
        category_labels.append("Uncategorized")
        category_data.append(uncategorized)

    return render(request, "supervisor/analytics.html", {
        "week_labels": week_labels, "week_data": week_data,
        "month_labels": month_labels, "month_data": month_data,
        "status_labels": list(status_counts.keys()),
        "status_data": list(status_counts.values()),
        "intern_labels": intern_labels, "intern_data": intern_data,
        "category_labels": category_labels, "category_data": category_data,
    })


# ---------------------------------------------------------------- reports

def report_queryset(request):
    tasks = visible_tasks_for(request.user)
    today = timezone.localdate()
    period = request.GET.get("period", "")
    if period == "daily":
        tasks = tasks.filter(date=today)
    elif period == "weekly":
        start, end = week_bounds(today)
        tasks = tasks.filter(date__range=(start, end))
    elif period == "monthly":
        tasks = tasks.filter(date__year=today.year, date__month=today.month)
    return filter_tasks(request, tasks).order_by("date", "start_time")


@login_required
def reports(request):
    tasks = report_queryset(request)
    total_hours = hours(tasks.aggregate(t=Coalesce(Sum("duration"), ZERO))["t"])
    return render(request, "reports/reports.html", {
        "tasks": tasks[:200],
        "task_count": tasks.count(),
        "total_hours": total_hours,
        "categories": categories_for(request.user),
        "interns": visible_interns_for(request.user) if request.user.is_management else None,
        "statuses": Task.Status.choices,
        "priorities": Task.Priority.choices,
    })


def report_rows(tasks, include_intern):
    header = ["Date", "Title", "Category", "Priority", "Status",
              "Start", "End", "Hours"]
    if include_intern:
        header.insert(1, "Intern")
    rows = [header]
    for t in tasks:
        row = [
            t.date.strftime("%Y-%m-%d"),
            t.title,
            t.category.name if t.category else "-",
            t.get_priority_display(),
            t.get_status_display(),
            t.start_time.strftime("%H:%M") if t.start_time else "-",
            t.end_time.strftime("%H:%M") if t.end_time else "-",
            t.duration_hours,
        ]
        if include_intern:
            row.insert(1, t.intern.get_full_name() or t.intern.username)
        rows.append(row)
    return rows


@login_required
def export_csv(request):
    tasks = report_queryset(request)
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="intrack_report.csv"'
    writer = csv.writer(response)
    for row in report_rows(tasks, request.user.is_management):
        writer.writerow(row)
    return response


@login_required
def export_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    tasks = report_queryset(request)
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"
    rows = report_rows(tasks, request.user.is_management)
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(color="FFFFFF", bold=True)
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for col in ws.columns:
        width = max(len(str(c.value or "")) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(width, 40)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="intrack_report.xlsx"'
    wb.save(response)
    return response


@login_required
def export_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    tasks = report_queryset(request)
    total = hours(tasks.aggregate(t=Coalesce(Sum("duration"), ZERO))["t"])

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="intrack_report.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4),
                            topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("InTrack — Task Report", styles["Title"]),
        Paragraph(
            f"Generated {timezone.localdate():%d %b %Y} · "
            f"{tasks.count()} task(s) · {total} hours total",
            styles["Normal"]),
        Spacer(1, 12),
    ]
    data = report_rows(tasks, request.user.is_management)
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F8FAFC")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(table)
    doc.build(elements)
    return response


# ---------------------------------------------------------------- chat

@login_required
def chat_inbox(request):
    conversations = (
        Conversation.objects.filter(participants=request.user)
        .annotate(
            unread=Count(
                "messages",
                filter=Q(messages__is_read=False) & ~Q(messages__sender=request.user),
            ),
            message_count=Count("messages"),
        )
        .prefetch_related("participants", "messages__sender")
    )
    rows = []
    for conv in conversations:
        other = conv.other_participant(request.user)
        last = conv.messages.all()
        last_msg = list(last)[-1] if last else None
        # Prefer last by created_at from prefetched cache
        if last:
            last_msg = max(last, key=lambda m: m.created_at)
        rows.append({
            "conversation": conv,
            "other": other,
            "last": last_msg,
            "unread": conv.unread,
        })

    start_form = StartChatForm(queryset=messageable_users_for(request.user))
    return render(request, "chat/inbox.html", {
        "rows": rows,
        "start_form": start_form,
    })


@login_required
def chat_start(request):
    """Start a chat via POST form or GET ?user=<id> from team/intern lists."""
    if request.method == "GET":
        user_id = request.GET.get("user")
        if not user_id:
            return redirect("chat_inbox")
        other = get_object_or_404(User, pk=user_id, is_active=True)
        if not can_message(request.user, other):
            messages.error(request, "You are not allowed to message that user.")
            return redirect("chat_inbox")
        conversation, _ = Conversation.between(request.user, other)
        return redirect("chat_thread", pk=conversation.pk)

    form = StartChatForm(request.POST, queryset=messageable_users_for(request.user))
    if form.is_valid():
        other = form.cleaned_data["recipient"]
        if not can_message(request.user, other):
            messages.error(request, "You are not allowed to message that user.")
            return redirect("chat_inbox")
        conversation, _ = Conversation.between(request.user, other)
        return redirect("chat_thread", pk=conversation.pk)
    messages.error(request, "Please choose a valid recipient.")
    return redirect("chat_inbox")


@login_required
def chat_thread(request, pk):
    conversation = get_object_or_404(
        Conversation.objects.prefetch_related("participants"), pk=pk)
    if not conversation.participants.filter(pk=request.user.pk).exists():
        return HttpResponseForbidden("Not your conversation.")

    other = conversation.other_participant(request.user)
    if other and not can_message(request.user, other):
        return HttpResponseForbidden("You are not allowed to message this user.")

    if request.method == "POST":
        form = MessageForm(request.POST)
        if form.is_valid():
            msg = Message.objects.create(
                conversation=conversation,
                sender=request.user,
                body=form.cleaned_data["body"].strip(),
            )
            conversation.save(update_fields=["updated_at"])
            if other:
                notify(other, f"Message from {request.user}",
                       msg.body[:80],
                       link=f"/chat/{conversation.pk}/")
            return redirect("chat_thread", pk=conversation.pk)
    else:
        form = MessageForm()

    conversation.messages.filter(is_read=False).exclude(sender=request.user).update(is_read=True)
    messages_qs = conversation.messages.select_related("sender")

    return render(request, "chat/thread.html", {
        "conversation": conversation,
        "other": other,
        "chat_messages": messages_qs,
        "form": form,
    })


# ---------------------------------------------------------------- profile / notifications

@login_required
def profile(request):
    if request.method == "POST":
        user_form = UserUpdateForm(request.POST, instance=request.user)
        profile_form = ProfileUpdateForm(
            request.POST, request.FILES, instance=request.user.profile)
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            messages.success(request, "Profile updated successfully.")
            return redirect("profile")
    else:
        user_form = UserUpdateForm(instance=request.user)
        profile_form = ProfileUpdateForm(instance=request.user.profile)
    return render(request, "accounts/profile.html", {
        "user_form": user_form,
        "profile_form": profile_form,
    })


@login_required
def notifications(request):
    items = request.user.notifications.all()[:50]
    request.user.notifications.filter(is_read=False).update(is_read=True)
    return render(request, "notifications/notification_list.html", {"items": items})
